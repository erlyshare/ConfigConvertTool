#! python
# -*- coding:utf-8 -*-

import os
import re
import sys
import json
import openpyxl
import color_print

from slpp.slpp import slpp as lua

# array数组模式下，各个栏的分布
# 1、2行是文档注释，不导出
ACMT_ROW = 3  # comment row 注释
ATPE_ROW = 4  # type    row 类型
ASRV_ROW = 5  # server  row 服务器
ACLT_ROW = 6  # client  row 客户端
AKEY_COL = 1  # key  column key所在列

# object(kv)模式下，各个栏的分布
OCMT_COL = 1  # comment column 注释
OTPE_COL = 2  # type    column 类型
OSRV_COL = 3  # server  column 服务器
OCLT_COL = 4  # client  column 客户端
OCTX_COL = 5  # content column 内容所在列

# 1、2行是文档注释，不导出
OFLG_ROW = 3  # flag    row    server client所在行

SRV_FLAG = "server"
CLT_FLAG = "client"

SHEET_FLAG_ROW = 3      # 3列1行表示是array 还是object 或者是不用导出的表
SHEET_FLAG_COL = 1      # 3列1行表示是array 还是object 或者是不用导出的表

# 用于标识表的类型 不是这两种就不会导出数据
ARRAY_FLAG  = "array"   # 标识表的是array类型的表
OBJECT_FLAG = "object"  # 标识表的是object类型的表

KEY_FLAG = "$key_"      # array表的 作为Key字段的前缀标识
KEY_FLAG_LEN = 5        # array表的 作为Key字段的前缀标识长度 用于分割字符串

# 支持的数据类型
# "int":1,"number":2,"int64":3,"string":4, "tuple":5, "list":6, "dict":7 为python原生数据格式
# "json":8,"lua":9 为json和lua的数据格式
# excel配置的时候 对应类型字段 填入对应数据类型的有效数据格式的数据就OK
TYPES = { "int":1,"number":2,"int64":3,"string":4, "tuple":5, "list":6, "dict":7, "json":8,"lua":9}

try:
    basestring
except NameError:
    basestring = str

# python3中没有uincode了，int能表示int64
try:
    long
except NameError:
    long = int

# python3中没有unicode了
try:
    unicode
except NameError:
    unicode = str

# 类型转换器
class ValueConverter(object):
    def __init__(self):
        pass

    # 在python中，字符串和unicode是不一样的。默认从excel读取的数据都是unicode。
    # str可以通过decode转换为unicode
    # ascii' codec can't encode characters
    # 这个函数在python3中没用
    def to_unicode_str( self,val ):
        if isinstance( val,str ) :
            return val
        elif isinstance( val,unicode ) :
            return val
        else :
            return str( val ).decode("utf8")

    def to_value(self,val_type,val):
        if "int" == val_type :
            return int( val )
        elif "int64" == val_type :
            return long( val )
        elif "number" == val_type :
            # 去除带小数时的小数点，100.0 ==>> 100
            # number就让它带小数点吧，不然强类型的配置无法正确识别出来
            # if long( val ) == float( val ) : return long( val )
            return float( val )
        elif "string" == val_type :
            return self.to_unicode_str( val )
        elif "json" == val_type :
            return json.loads( val )
        elif "lua" == val_type :
            return lua.decode( val )
        elif "tuple" == val_type :
            return  tuple(eval(val))
        elif "list" == val_type :
            return list(eval(val))
        elif "dict" == val_type :
            return dict(eval(val))
        else :
            self.raise_error( "invalid type",value )

class Sheet(object):

    def __init__(self,base_name,wb_sheet,srv_writer,clt_writer):

        self.types  = []        # 记录各列字段的类型

        self.srv_writer = srv_writer
        self.clt_writer = clt_writer

        self.srv_fields = []    #服务端各列字段名
        self.clt_fields = []    #客户端各列字段名
        self.srv_comment = {}   # 服务器字段注释
        self.clt_comment = {}   # 客户端字段注释

        self.srv_keys = {}      # 服务器key
        self.clt_keys = {}      # 客户端key


        self.converter = ValueConverter()

        self.wb_sheet  = wb_sheet
        self.base_name = base_name

        match_list = re.findall("[a-zA-Z0-9_]+$",base_name)
        if None == match_list or 1 != len(match_list):
            Exception( base_name,"not a legal file name" )

        self.base_file_name = match_list[0]

        # 记录出错时的行列，方便策划定位问题
        self.error_row = 0
        self.error_col = 0

    # 记录出错位置
    def mark_error_pos(self,row,col):
        if row > 0 : self.error_row = row
        if col > 0 : self.error_col = col

    # 发起一个解析错误
    def raise_error(self,what,val):
        excel_info = format("DOC:%s,SHEET:%s,ROW:%d,COLUMN:%d" % \
            (self.base_name,self.wb_sheet.title,self.error_row,self.error_col))
        raise Exception( what,val,excel_info )

    def to_value(self,val_type,val):
        try:
            return self.converter.to_value(val_type,val)
        except Exception :
            t, e = sys.exc_info()[:2]
            self.raise_error( "ConverError",e )

    # 解析一个表格
    def decode_sheet(self):
        wb_sheet = self.wb_sheet

        self.decode_type ()
        self.decode_field()
        self.decode_ctx  ()

        color_print.printGreen( "        covert successfully... sheet name -> %s \n" % wb_sheet.title )
        return True

    # 写入配置到文件
    def write_one_file(self,ctx,base_path,writer, keys_list, comment_text):
        # 有些配置可能只导出客户端或只导出服务器
        if not any(ctx) :
            return

        write_file_name = self.base_file_name
        if self.wb_sheet.title.find("+") >= 0 or self.wb_sheet.title.find("-") >= 0 :
            match_list = re.findall("[a-zA-Z0-9_]+$", self.wb_sheet.title)
            if 1 == len(match_list):
                write_file_name = write_file_name + '_' + match_list[0]

        wt = writer(self.base_name,self.wb_sheet.title, write_file_name, keys_list, comment_text)
        ctx = wt.context( ctx )
        suffix = wt.suffix()
        if None != ctx :
            path = base_path + write_file_name + suffix

            #必须为wb，不然无法写入utf-8
            file = open( path, 'wb' )
            file.write( ctx.encode( "utf-8" ) )
            file.close()

    # 分别写入到服务端、客户端的配置文件
    def write_files(self,srv_path,clt_path):
        if None != srv_path and None != self.srv_writer :
            self.write_one_file( self.srv_ctx,srv_path,self.srv_writer, self.srv_keys, self.srv_comment)
        if None != clt_path and None != self.clt_writer :
            self.write_one_file( self.clt_ctx,clt_path,self.clt_writer, self.clt_keys, self.clt_comment )

# 导出数组类型配置，A1格子的内容有array标识
class ArraySheet(Sheet):

    def __init__(self,base_name,wb_sheet,srv_writer,clt_writer):
        # 记录导出各行的内容
        self.srv_ctx = []
        self.clt_ctx = []

        super( ArraySheet, self ).__init__(
            base_name,wb_sheet,srv_writer,clt_writer )

    # 解析各列的类型(string、number...)
    def decode_type(self):
        # 第一列没数据，类型可以不填，默认为None，但是这里要占个位
        self.types.append( None )

        for col_idx in range( AKEY_COL + 1,self.wb_sheet.max_column + 1 ):
            self.mark_error_pos(ATPE_ROW,col_idx)
            value = self.wb_sheet.cell( row = ATPE_ROW, column = col_idx ).value

            # 单元格为空的时候，wb_sheet.cell(row=1, column=2).value == None
            # 类型那一行必须连续，空白表示后面的数据都不导出了
            if value == None : 
                break
            if value not in TYPES :
                self.raise_error( "invalid type",value )

            self.types.append( value )

    # 解析客户端、服务器的字段名(server、client)那两行
    def decode_one_field(self,fields,row_index, keys):
        key_index = 1
        for col_index in range( AKEY_COL,len( self.types ) + 1 ):
            value = self.wb_sheet.cell(
                row = row_index, column = col_index ).value
            # 对于array类型的表 一条数据可以有多个值作为KEY 需要处理一下
            if None != value and value.find(KEY_FLAG) >= 0 :
                value = value[KEY_FLAG_LEN:]
                keys[value] = key_index
                key_index = key_index + 1

            # 对于不需要导出的field，可以为空。即value为None
            fields.append( value )

    # 解析注释那一行
    def decode_one_comment(self,fields,row_index, key_index):
        for col_index in range( AKEY_COL + 1,len( self.types ) + 1 ):
            value = self.wb_sheet.cell(
                row = row_index, column = col_index ).value

            key = self.wb_sheet.cell(
                    row = key_index, column = col_index ).value

            if None == key :
                continue

            # 对于不需要导出的field，可以为空。即value为None
            if key.find(KEY_FLAG) >= 0 :
                key = key[KEY_FLAG_LEN:]
                fields[key] = value
            else :
                fields[key] = value

    # 导出客户端、服务端字段名(server、client)那一列
    def decode_field(self):
        self.decode_one_field( self.srv_fields,ASRV_ROW, self.srv_keys )
        self.decode_one_field( self.clt_fields,ACLT_ROW, self.clt_keys )
        # 导出服务器和客户端对用字段的注释， 导表的时候可能用到
        self.decode_one_comment( self.srv_comment, ACMT_ROW,  ASRV_ROW)
        self.decode_one_comment( self.clt_comment, ACMT_ROW,  ACLT_ROW)

    # 解析出一个格子的内容
    def decode_cell(self,row_idx,col_idx):
        value = self.wb_sheet.cell( row = row_idx, column = col_idx ).value
        if None == value:
            return None

        # 类型是从0下标开始，但是excel的第一列从1开始
        self.mark_error_pos( row_idx,col_idx )
        return self.to_value( self.types[col_idx - 1],value )

    # 解析出一行的内容
    def decode_row(self,row_idx):
        srv_row = {}
        clt_row = {}

        # 第一列没数据，从第二列开始解析
        for col_idx in range( AKEY_COL + 1,len( self.types ) + 1 ):
            value = self.decode_cell( row_idx,col_idx )
            if None == value : continue

            srv_key = self.srv_fields[col_idx - 1]
            clt_key = self.clt_fields[col_idx - 1]


            if srv_key :
                srv_row[srv_key] = value
            if clt_key :
                clt_row[clt_key] = value
        return srv_row,clt_row # 返回一个tuple

    # 解析导出的内容
    def decode_ctx(self):
        for row_idx in range( ACLT_ROW + 1,self.wb_sheet.max_row + 1 ):
            srv_row,clt_row = self.decode_row( row_idx )

            if row_idx > 57 :
                break

            # 不为空才追加
            if any( srv_row ) :
                self.srv_ctx.append( srv_row )
            if any( clt_row ) :
                self.clt_ctx.append( clt_row )

# 导出object类型的结构，A1格子有object标识
class ObjectSheet(Sheet):

    def __init__(self,base_name,wb_sheet,srv_writer,clt_writer):
        # 记录导出各行的内容
        self.srv_ctx = {}
        self.clt_ctx = {}

        super( ObjectSheet, self ).__init__(
            base_name,wb_sheet,srv_writer,clt_writer )

    # 解析各字段的类型
    def decode_type(self):
        for row_idx in range( OFLG_ROW + 1,self.wb_sheet.max_row + 1 ):
            self.mark_error_pos( row_idx,OTPE_COL)
            value = self.wb_sheet.cell( row = row_idx, column = OTPE_COL ).value

            # 类型必须连续，遇到空则认为后续数据不再导出
            if value == None :
                break
            if value not in TYPES :
                self.raise_error( "invalid type",value )

            self.types.append( value )

    # 导出客户端、服务端字段名(server、client)那一列
    def decode_one_field(self,fields,col_idx):
        for row_idx in range( OFLG_ROW + 1,len( self.types ) + OFLG_ROW + 1 ):
            value = self.wb_sheet.cell( row = row_idx, column = col_idx ).value

            # 对于不需要导出的field，可以为空。即value为None
            fields.append( value )

    # 解析注释那一行
    def decode_one_comment(self,fields,col_idx, key_index):
        for row_idx in range( OFLG_ROW + 1,len( self.types ) + OFLG_ROW + 1 ):
            value = self.wb_sheet.cell( row = row_idx, column = col_idx ).value

            key = self.wb_sheet.cell( row = row_idx, column = key_index ).value

            if None == value:
                continue
            # 导出存在值的fields的注释
            fields[key] =  value

    # 导出客户端、服务端字段名(server、client)那一列
    def decode_field(self):
        self.decode_one_field( self.srv_fields,OSRV_COL )
        self.decode_one_field( self.clt_fields,OCLT_COL )
        # 导出服务器和客户端对用字段的注释， 导表的时候可能用到
        self.decode_one_comment( self.srv_comment, OCMT_COL, OSRV_COL)
        self.decode_one_comment( self.clt_comment, OCMT_COL, OCLT_COL )

    # 解析一个单元格内容
    def decode_cell(self,row_idx):
        value = self.wb_sheet.cell( row = row_idx, column = OCTX_COL ).value
        if None == value : return None

        # 在object的结构中，数据是从第二行开始的，所以types的下标偏移2
        self.mark_error_pos( row_idx,OCTX_COL )
        return self.to_value( self.types[row_idx - OFLG_ROW - 1],value )

    # 解析表格的所有内容
    def decode_ctx(self):
        for row_idx in range( OFLG_ROW + 1,len( self.types ) + OFLG_ROW + 1 ):
            value = self.decode_cell( row_idx )

            if row_idx > 54 :
                break

            if None == value :
                continue

            srv_key = self.srv_fields[row_idx - OFLG_ROW - 1]
            clt_key = self.clt_fields[row_idx - OFLG_ROW - 1]

            if srv_key : self.srv_ctx[srv_key] = value
            if clt_key : self.clt_ctx[clt_key] = value

class ExcelDoc:

    def __init__(self, file,abspath):
        self.file = file
        self.abspath = abspath

    # 是否需要解析
    # 返回解析的对象类型
    def need_decode(self,wb_sheet):
        sheet_val = wb_sheet.cell(
            row = SHEET_FLAG_ROW, column = SHEET_FLAG_COL ).value

        sheeter = None
        srv_value = None
        clt_value = None
        if ARRAY_FLAG == sheet_val :
            if wb_sheet.max_row <= ACLT_ROW or wb_sheet.max_column <= AKEY_COL:
                return None

            sheeter = ArraySheet
            srv_value = wb_sheet.cell( row = ASRV_ROW, column = AKEY_COL ).value
            clt_value = wb_sheet.cell( row = ACLT_ROW, column = AKEY_COL ).value
        elif OBJECT_FLAG == sheet_val :
            sheeter = ObjectSheet
            srv_value = wb_sheet.cell( row = OFLG_ROW, column = OSRV_COL ).value
            clt_value = wb_sheet.cell( row = OFLG_ROW, column = OCLT_COL ).value
        else :
            return None

        # 没有这两个标识就不是配置表。可能是策划的一些备注说明
        if SRV_FLAG != srv_value or CLT_FLAG != clt_value :
            return None
        return sheeter

    def decode(self,srv_path,clt_path,srv_writer,clt_writer):
        color_print.printYellow( "    start covert: %s \n" % self.file.ljust(44, "*") )
        base_name = os.path.splitext( self.file )[0]
        wb = openpyxl.load_workbook( self.abspath )

        for wb_sheet in wb.worksheets:
            Sheeter = self.need_decode( wb_sheet )
            if None == Sheeter :
                color_print.printPink( "        covert skip........... sheet name -> %s\n" % wb_sheet.title )
                continue

            sheet = Sheeter( base_name,wb_sheet,srv_writer,clt_writer )
            if sheet.decode_sheet() :
                sheet.write_files( srv_path,clt_path )